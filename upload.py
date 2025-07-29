import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

def update_excel_date(file_path, searchitem,colName,new_value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict= {}

    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i,column=j).value == searchitem:
                Dict["row"]= i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)

file_path = "C:\\Users\\komma\\Downloads\\download.xlsx"
fruit_name= "Apple"
new_value = "999"
driver = webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.implicitly_wait(3)
driver.find_element(By.ID,"downloadButton").click()

update_excel_date(file_path ,fruit_name,"price" ,new_value)



driver.find_element(By.XPATH,"//input[@type='file']").send_keys(file_path)
Wait = WebDriverWait(driver,3)
Wait.until(expected_conditions.visibility_of_element_located((By.XPATH,"//div[@class='Toastify__toast-body']/div[2]")))
print(driver.find_element(By.XPATH,"//div[@class='Toastify__toast-body']/div[2]").text)

price_column =driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']/div").text
print(actual_price)


